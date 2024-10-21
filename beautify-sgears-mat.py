import argparse
import pandas as pd
from openpyxl import Workbook


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Beautify Silent Gear's Material Dump")
    parser.add_argument(
        "-i",
        "--input",
        required=False,
        type=str,
        help="Path to Silent Gear's Materials TSV dump",
        default="./DELETELATER/material_export.tsv",
    )
    args = parser.parse_args()

    return args.input


def main():
    input_file = parse_arguments()

    output_file = "./materials_beautified.csv"

    required_fieldnames = [
        # Random
        "Name",  # df4343
        "Type",  # ee3f72
        "Tier",  # f04ca3
        "Rarity",  # e464d1
        "Enchantment Value",  # c87ff9
        "Charging Value",  # 9d9cff
        # Durability and Repair #72b3ff
        "Durability",
        "Armor Durability",
        "Repair Efficiency",
        "Repair Bonus",
        # Toothpick #3dc3ff
        "Harvest Speed",
        # Long Hands #00d2ff
        "Reach Distance",
        # Attack #00def8
        "Attack Damage",
        "Attack Speed",
        "Attack Reach",
        "Magic Damage",
        # Ranged and Projectile #00eae3
        "Ranged Damage",
        "Ranged Speed",
        "Projectile Speed",
        "Projectile Accuracy",
        # Armor #00f5c5
        "Armor",
        "Armor Toughness",
        "Knockback Resistance",
        "Magic Armor",
        # Traits #70faa7
        "Traits",
    ]

    # Read the materials TSV file
    materials = pd.read_csv(input_file, sep="\t")

    # # DEBUG
    # # Print the first few rows to check the data
    # print("Initial DataFrame:")
    # print(materials.head())
    #
    # # Check for NaN in 'Parent' and 'ID' values
    # print("Checking filtering conditions:")
    # print("NaN in 'Parent':", materials["Parent"].isna().sum())
    # print("Non-matching 'ID':", (materials["ID"] == "silentgear:example").sum())

    # Filter and sort materials DataFrame
    materials = (
        (
            materials[
                materials["Parent"].isna() & (materials["ID"] != "silentgear:example")
            ][required_fieldnames]
        )
        .sort_values(["Type", "Tier"])
        .reset_index(drop=True)
    )

    # Adding empty Blank Rows when value in "Type" column changes?
    mask = materials["Type"].ne(materials["Type"].shift(-1))

    print(mask.head(12))

    empty_df = pd.DataFrame("", index=mask.index[mask] + 0.5, columns=materials.columns)

    materials = (
        pd.concat([materials, empty_df]).sort_index().reset_index(drop=True).iloc[:-1]
    )

    # Output the filtered DataFrame
    print(materials.head(12))

    # wb = Workbook()
    #
    # general_ws = wb.create_sheet("General", 0)
    # tools_ws = wb.create_sheet("Tools", 1)
    # weapons_ws = wb.create_sheet("Weapons", 2)
    # armor_ws = wb.create_sheet("Armor", 3)
    #
    # general_ws.append(required_fieldnames)
    #
    # for row in materials:
    #     general_ws.append(row.values.tolist())

    # wb.save("materials_beautified.xlsx")

    # Uncomment to save the beautified DataFrame to CSV
    materials.to_csv(output_file, sep="\t", index=False)


if __name__ == "__main__":
    main()
